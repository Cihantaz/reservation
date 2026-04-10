import openpyxl

wb = openpyxl.load_workbook('test_schedule.xlsx')
ws = wb.active
# 4. satıra yeni veri ekle
ws['A4'] = 'MATH200'
ws['B4'] = 'DMF-114'
ws['C4'] = 'W3'

wb.save('test_schedule.xlsx')
print('Excel dosyası güncellendi (3. örnek eklendi)')
