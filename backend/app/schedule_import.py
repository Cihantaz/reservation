import csv
import io
import re
from decimal import Decimal
from dataclasses import dataclass
from typing import Dict, Optional

import pandas as pd
from openpyxl import load_workbook


VALID_WEEKDAYS = {"M", "T", "W", "TH", "F"}


@dataclass(frozen=True)
class ParsedRow:
    course_code: str
    room_name: str
    weekday: str  # M/T/W/TH/F
    slot: int  # 1..12


def _split_csv(cell: str) -> list[str]:
    if cell is None:
        return []
    parts = [p.strip() for p in str(cell).split(",")]
    return [p for p in parts if p]


_TIME_TOKEN_RE = re.compile(r"^(TH|M|T|W|F)\s*([1-9]|1[0-2])$", re.IGNORECASE)

_ROOM_TOKEN_RE = re.compile(r"^([A-ZÇĞİÖŞÜ]+)\s*-?\s*(\d+)$", re.IGNORECASE)


_MOJIBAKE_REPLACEMENTS = {
    "Ã‡": "Ç",
    "Ã§": "ç",
    "Ä": "Ğ",
    "ÄŸ": "ğ",
    "Ä°": "İ",
    "Ä±": "ı",
    "Ã–": "Ö",
    "Ã¶": "ö",
    "Å": "Ş",
    "ÅŸ": "ş",
    "Ãœ": "Ü",
    "Ã¼": "ü",
    "Â": "",
    "�": "",
}


def _repair_text(value: object) -> str:
    text = "" if value is None else str(value)
    for broken, fixed in _MOJIBAKE_REPLACEMENTS.items():
        text = text.replace(broken, fixed)
    return text.strip()


def _normalize_label(value: object) -> str:
    text = _repair_text(value).lower()
    text = text.replace("i̇", "i")
    for src, dst in {
        "ı": "i",
        "ö": "o",
        "ü": "u",
        "ş": "s",
        "ğ": "g",
        "ç": "c",
    }.items():
        text = text.replace(src, dst)
    return re.sub(r"[^a-z0-9()]+", "", text)


def _find_col(columns: list[str], candidates: list[str]) -> Optional[str]:
    colset = {_normalize_label(c): str(c) for c in columns if _normalize_label(c)}
    for cand in candidates:
        match = colset.get(_normalize_label(cand))
        if match is not None:
            return match
    return None


def _find_col_index(columns: list[str], candidates: list[str]) -> Optional[int]:
    target = _find_col(columns, candidates)
    if target is None:
        return None
    for idx, col in enumerate(columns):
        if str(col) == target:
            return idx
    return None


def _error(row: Optional[int], message: str, detail: str = "") -> dict:
    payload = {"row": row, "message": message}
    if detail:
        payload["detail"] = detail
    return payload


def _columns_detail(columns: list[object]) -> str:
    names = [_repair_text(col) for col in columns if _repair_text(col)]
    return f"Bulunan kolonlar: {', '.join(names)}" if names else "Dosyada okunabilen kolon bulunamadı."


def _row_detail(values: dict[str, object]) -> str:
    parts = []
    for key, value in values.items():
        text = _repair_text(value)
        if text:
            parts.append(f"{key}='{text}'")
    return "Veri: " + ", ".join(parts) if parts else ""


def _cell_text(cell) -> str:
    value = cell.value
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if isinstance(value, Decimal) and value == int(value):
        value = int(value)
    return _repair_text(value)


def _room_number_text(cell) -> str:
    value = cell.value
    if value is None:
        return ""

    if isinstance(value, Decimal):
        value = int(value) if value == int(value) else float(value)

    if isinstance(value, int):
        fmt = str(cell.number_format or "").strip()
        if re.fullmatch(r"0+", fmt):
            return f"{value:0{len(fmt)}d}"
        return str(value)

    if isinstance(value, float):
        if value.is_integer():
            int_value = int(value)
            fmt = str(cell.number_format or "").strip()
            if re.fullmatch(r"0+", fmt):
                return f"{int_value:0{len(fmt)}d}"
            return str(int_value)
        return _repair_text(value)

    return _repair_text(value)


def derive_exam_capacity(building: str, class_capacity: int) -> int:
    if str(building).strip().upper() == "DK":
        return 25
    return class_capacity // 2


def _normalize_room_code(token: str, available_rooms: Optional[Dict[str, str]] = None) -> list[str]:
    """
    Excel'deki sınıf hücreleri normalize et ve DB'deki sınıflarla eşle:
    
    INPUT FORMATS (Excel'den gelebilecek):
      - A203 → A-203 (format standardization)
      - DMF-114 → DMF-114 (already normalized)
      - "A413 A317" → [A-413, A-317] (multiple classes same time)
      - dmf-102 → DMF-102 (case normalization via fuzzy matching)
    
    MATCHING STRATEGY (FUZZY & CASE-INSENSITIVE):
      1. Reformat input: A203 → A-203 (standard Bina-DerslikNum format)
      2. Check exact match in DB (A-203 == A-203)
      3. Fuzzy match: case-insensitive & hyphen-tolerant (dmf-102 matches DMF-102)
      4. If not found: return normalized form (error will be caught in preview)
    
    available_rooms: {room_name: room_id} dict. If provided, enables fuzzy matching.
    """
    if token is None:
        return []
    token = str(token).strip()
    if not token:
        return []
    
    # Split multiple rooms (space or comma separated)
    parts = [p.strip() for p in re.split(r"[\s,]+", token) if p.strip()]
    out: list[str] = []
    
    for p in parts:
        normalized = None
        
        # STEP 1: Standardize format using regex (A203 → A-203)
        m = _ROOM_TOKEN_RE.match(p.replace(" ", ""))
        if m:
            b = m.group(1).upper()  # Building (DMF, A, B, etc.)
            num = m.group(2)         # Room number (101, 203, etc.)
            normalized = f"{b}-{num}"
        else:
            # Fallback if no regex match (might be already normalized)
            normalized = p.strip().upper()
        
        # STEP 2: Fuzzy matching with DB (if available_rooms provided)
        if available_rooms:
            # 2a. EXACT MATCH first (A-203 == A-203)
            exact_match = available_rooms.get(normalized)
            if exact_match:
                out.append(normalized)
                continue
            
            # 2b. FUZZY MATCH: Case-insensitive & hyphen-tolerant
            # Example: dmf-102 matches DMF-102 even with different case/hyphen
            for db_name in available_rooms.keys():
                if (db_name.upper() == normalized or 
                    db_name.replace("-", "").upper() == normalized.replace("-", "")):
                    out.append(db_name)  # Use actual DB name for consistency
                    break
            else:
                # Not found → append normalized form (preview validation will flag error)
                out.append(normalized)
        else:
            # No DB provided, just append normalized form
            out.append(normalized)
    
    return out if out else [token.strip()]


def parse_schedule_excel(content: bytes, available_rooms: Optional[Dict[str, str]] = None) -> dict:
    """
    Excel kolonları:
      - Ders Kodu
      - Sınıf(lar)
      - Ders Saati

    Kurallar:
      - Sınıf(lar) ve Ders Saati virgülle ayrılmışsa, array'ler ZIP edilerek eşlenir.
        Örn: A203, A204 | T4, T5  => (A203,T4) + (A204,T5)
      - Eğer biri 1 eleman, diğeri N eleman ise 1 eleman tümüne uygulanır.
      - Ders Saati token formatı: M1..M12, T1..T12, W1..W12, TH1..TH12, F1..F12
      - available_rooms: {room_name: room_id} dict, fuzzy matching için kullanılır
    """
    try:
        df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        return {
            "ok": False,
            "errors": [{"row": None, "message": f"Excel parse hatası: {str(e)}"}],
            "items": [],
            "total_items": 0,
        }

    col_course = _find_col(list(df.columns), ["Ders Kodu"])
    col_rooms = _find_col(list(df.columns), ["Sınıf(lar)", "S\u0131n\u0131f(lar)", "S\uFFFdn\uFFFDf(lar)"])
    col_time = _find_col(list(df.columns), ["Ders Saati"])
    col_email = _find_col(list(df.columns), ["E-Posta", "Eposta", "E-mail", "Email"])

    missing = []
    if not col_course:
        missing.append("Ders Kodu")
    if not col_rooms:
        missing.append("Sınıf(lar)")
    if not col_time:
        missing.append("Ders Saati")
    if missing:
        return {
            "ok": False,
            "errors": [_error(None, f"Eksik kolon(lar): {', '.join(missing)}", _columns_detail(list(df.columns)))],
            "items": [],
            "total_items": 0,
        }

    items: list[dict] = []
    errors: list[dict] = []

    for idx, row in df.iterrows():
        excel_row = int(idx) + 2  # header=1
        course_code = str(row.get(col_course, "")).strip()
        rooms_raw = str(row.get(col_rooms, "")).strip()
        times_raw = str(row.get(col_time, "")).strip()
        actor_email = str(row.get(col_email, "")).strip().lower() if col_email else ""
        detail = _row_detail(
            {
                "Ders Kodu": course_code,
                "Sınıf(lar)": rooms_raw,
                "Ders Saati": times_raw,
                "E-Posta": actor_email,
            }
        )

        if not course_code:
            errors.append(_error(excel_row, "Ders Kodu boş olamaz.", detail))
            continue
        if not rooms_raw:
            errors.append(_error(excel_row, "Sınıf(lar) boş olamaz.", detail))
            continue
        if not times_raw:
            errors.append(_error(excel_row, "Ders Saati boş olamaz.", detail))
            continue

        rooms = _split_csv(rooms_raw)
        times = _split_csv(times_raw)

        if not rooms or not times:
            errors.append(_error(excel_row, "Sınıf(lar) veya Ders Saati parse edilemedi.", detail))
            continue

        pairs: list[tuple[str, str]] = []
        if len(rooms) == len(times):
            pairs = list(zip(rooms, times))
        elif len(rooms) == 1 and len(times) > 1:
            pairs = [(rooms[0], t) for t in times]
        elif len(times) == 1 and len(rooms) > 1:
            pairs = [(r, times[0]) for r in rooms]
        else:
            errors.append(
                _error(
                    excel_row,
                    f"Sınıf(lar) ve Ders Saati eşleşmiyor (zip edilemedi). Sınıf sayısı={len(rooms)}, saat sayısı={len(times)}.",
                    detail,
                )
            )
            continue

        for (room_name, token) in pairs:
            m = _TIME_TOKEN_RE.match(token.replace(" ", ""))
            if not m:
                errors.append(_error(excel_row, f"Ders Saati formatı hatalı: '{token}' (örn: T4, TH10).", detail))
                continue
            wd = m.group(1).upper()
            slot = int(m.group(2))
            if wd not in VALID_WEEKDAYS:
                errors.append(_error(excel_row, f"Geçersiz gün: '{wd}'", detail))
                continue

            # "A413 A317" gibi çoklu sınıf varsa, aynı saat için ayrı kayıt üret
            room_codes = _normalize_room_code(room_name, available_rooms=available_rooms)
            for rc in room_codes:
                items.append(
                    {
                        "course_code": course_code,
                        "room_name": rc,
                        "weekday": wd,
                        "slot": slot,
                        "source_row": excel_row,
                        "actor_email": actor_email,
                    }
                )

    return {"ok": len(errors) == 0, "errors": errors, "items": items, "total_items": len(items)}


def parse_rooms_excel(content: bytes) -> dict:
    """
    Excel kolonları:
      - Bina
      - Derslik Numarası
      - Özellik
      - Kapasite

    İş kuralları:
      - room_code = f\"{Bina}-{Derslik Numarası}\" (örn DMF-114) => Room.name
      - Sınav Kapasitesi = floor(Kapasite / 2)
      - DK binasında sınav kapasitesi sabit 25
      - UPSERT: room_code varsa update, yoksa insert (save endpoint’inde)
    """
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        return {
            "ok": False,
            "errors": [_error(None, f"Excel parse hatası: {str(e)}")],
            "items": [],
            "total_items": 0,
        }

    ws = wb[wb.sheetnames[0]]
    headers = [_cell_text(ws.cell(1, col_idx)) for col_idx in range(1, ws.max_column + 1)]

    col_building = _find_col_index(headers, ["Bina"])
    col_number = _find_col_index(headers, ["Derslik Numarası", "Derslik Numarasi", "Derslik"])
    col_feature = _find_col_index(headers, ["Özellik", "Ozellik"])
    col_capacity = _find_col_index(headers, ["Kapasite"])

    missing = []
    if col_building is None:
        missing.append("Bina")
    if col_number is None:
        missing.append("Derslik")
    if col_feature is None:
        missing.append("Özellik")
    if col_capacity is None:
        missing.append("Kapasite")
    if missing:
        return {
            "ok": False,
            "errors": [_error(None, f"Eksik kolon(lar): {', '.join(missing)}", _columns_detail(headers))],
            "items": [],
            "total_items": 0,
        }

    items: list[dict] = []
    errors: list[dict] = []

    for excel_row in range(2, ws.max_row + 1):
        building_cell = ws.cell(excel_row, col_building + 1)
        number_cell = ws.cell(excel_row, col_number + 1)
        feature_cell = ws.cell(excel_row, col_feature + 1)
        capacity_cell = ws.cell(excel_row, col_capacity + 1)

        building = _cell_text(building_cell).upper()
        number = _room_number_text(number_cell)
        feature = _cell_text(feature_cell)
        cap_raw = capacity_cell.value
        detail = _row_detail(
            {
                "Bina": building,
                "Derslik": number,
                "Özellik": feature,
                "Kapasite": cap_raw,
            }
        )

        if not building:
            errors.append(_error(excel_row, "Bina boş olamaz.", detail))
            continue
        if not number:
            errors.append(_error(excel_row, "Derslik Numarası boş olamaz.", detail))
            continue

        try:
            cap = int(cap_raw)
        except Exception:
            errors.append(_error(excel_row, f"Kapasite sayısal olmalı: '{cap_raw}'", detail))
            continue

        if cap < 0:
            errors.append(_error(excel_row, "Kapasite negatif olamaz.", detail))
            continue

        room_code = f"{building}-{number}"
        exam_capacity = derive_exam_capacity(building, cap)
        items.append(
            {
                "room_code": room_code,
                "building": building,
                "room_number": number,
                "feature": feature,
                "class_capacity": cap,
                "exam_capacity": exam_capacity,
                "source_row": excel_row,
            }
        )

    return {"ok": len(errors) == 0, "errors": errors, "items": items, "total_items": len(items)}


def parse_schedule_csv(content: bytes, available_rooms: Optional[Dict[str, str]] = None) -> dict:
    """
    CSV/TSV format'da ders takvimi parse et.
    Tab (\\t) veya virgül (,) ile ayrılmış veri kabul eder.
    
    Format:
      Ders Kodu\tSınıf(lar)\tDers Saati\t[E-Posta (opsiyonel)]
      MATH101\tA-203\tM4
      PHYS201\tB-114, B-115\tT5, T6
      
    available_rooms: {room_name: room_id} dict, fuzzy matching için kullanılır
    """
    try:
        content_str = content.decode('utf-8', errors='replace').strip()
        
        # Delimiter detect: tab veya virgül
        lines = content_str.split('\n')
        if not lines:
            return {
                "ok": False,
                "errors": [{"row": None, "message": "CSV dosyası boş."}],
                "items": [],
                "total_items": 0,
            }
        
        # CSV detect
        delimiter = '\t' if '\t' in lines[0] else ','
        
        # CSV parse
        reader = csv.reader(io.StringIO(content_str), delimiter=delimiter)
        rows = list(reader)
        
        if not rows:
            return {
                "ok": False,
                "errors": [{"row": None, "message": "CSV parse hatası."}],
                "items": [],
                "total_items": 0,
            }
        
        # Header
        headers = [h.strip().lower() for h in rows[0]]
        
        col_idx_course = None
        col_idx_rooms = None
        col_idx_time = None
        col_idx_email = None
        
        for i, h in enumerate(headers):
            if "ders kodu" in h:
                col_idx_course = i
            elif "sınıf" in h or "sinif" in h:
                col_idx_rooms = i
            elif "ders saati" in h or "saat" in h:
                col_idx_time = i
            elif "posta" in h or "mail" in h:
                col_idx_email = i
        
        missing = []
        if col_idx_course is None:
            missing.append("Ders Kodu")
        if col_idx_rooms is None:
            missing.append("Sınıf(lar)")
        if col_idx_time is None:
            missing.append("Ders Saati")
        
        if missing:
            return {
                "ok": False,
                "errors": [{"row": None, "message": f"Eksik kolon(lar): {', '.join(missing)}"}],
                "items": [],
                "total_items": 0,
            }
        
        items: list[dict] = []
        errors: list[dict] = []
        
        for row_idx, row in enumerate(rows[1:], start=2):
            # Padding
            while len(row) <= max(col_idx_course, col_idx_rooms, col_idx_time):
                row.append("")
            
            course_code = row[col_idx_course].strip()
            rooms_raw = row[col_idx_rooms].strip()
            times_raw = row[col_idx_time].strip()
            actor_email = row[col_idx_email].strip().lower() if col_idx_email and col_idx_email < len(row) else ""
            
            if not course_code:
                errors.append({"row": row_idx, "message": "Ders Kodu boş olamaz."})
                continue
            if not rooms_raw:
                errors.append({"row": row_idx, "message": "Sınıf(lar) boş olamaz."})
                continue
            if not times_raw:
                errors.append({"row": row_idx, "message": "Ders Saati boş olamaz."})
                continue
            
            rooms = _split_csv(rooms_raw)
            times = _split_csv(times_raw)
            
            if not rooms or not times:
                errors.append({"row": row_idx, "message": "Sınıf(lar) veya Ders Saati parse edilemedi."})
                continue
            
            # Pairing logic
            pairs: list[tuple[str, str]] = []
            if len(rooms) == len(times):
                pairs = list(zip(rooms, times))
            elif len(rooms) == 1 and len(times) > 1:
                pairs = [(rooms[0], t) for t in times]
            elif len(times) == 1 and len(rooms) > 1:
                pairs = [(r, times[0]) for r in rooms]
            else:
                errors.append({
                    "row": row_idx,
                    "message": f"Sınıf(lar) ve Ders Saati eşleşmiyor. Sınıf={len(rooms)}, saat={len(times)}.",
                })
                continue
            
            for (room_name, token) in pairs:
                m = _TIME_TOKEN_RE.match(token.replace(" ", ""))
                if not m:
                    errors.append({"row": row_idx, "message": f"Ders Saati formatı hatalı: '{token}' (örn: M4, TH10)."})
                    continue
                
                wd = m.group(1).upper()
                slot = int(m.group(2))
                
                if wd not in VALID_WEEKDAYS:
                    errors.append({"row": row_idx, "message": f"Geçersiz gün: '{wd}'"})
                    continue
                
                room_codes = _normalize_room_code(room_name, available_rooms=available_rooms)
                for rc in room_codes:
                    items.append({
                        "course_code": course_code,
                        "room_name": rc,
                        "weekday": wd,
                        "slot": slot,
                        "source_row": row_idx,
                        "actor_email": actor_email,
                    })
        
        return {"ok": len(errors) == 0, "errors": errors, "items": items, "total_items": len(items)}
    
    except Exception as e:
        return {
            "ok": False,
            "errors": [{"row": None, "message": f"CSV parse hatası: {str(e)}"}],
            "items": [],
            "total_items": 0,
        }

