import openpyxl

wb = openpyxl.load_workbook('test_schedule.xlsx')
ws = wb.active
print('Rows:')
for i in range(1, ws.max_row + 1):
    a = ws[f'A{i}'].value
    b = ws[f'B{i}'].value
    c = ws[f'C{i}'].value
    print(f'{i}: {a} | {b} | {c}')
